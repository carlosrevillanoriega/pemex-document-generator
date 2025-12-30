from flask import Flask, request, jsonify, send_from_directory
import os
import pandas as pd
import sys
import subprocess
import json
from datetime import datetime
import re
from werkzeug.utils import secure_filename
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

app = Flask(__name__, static_folder='static', static_url_path='')

# Configuración para subir archivos
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Crear carpeta de uploads si no existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/generar_documentos', methods=['POST'])
def generar_documentos():
    try:
        # Verificar si se enviaron archivos
        if 'archivo' not in request.files or 'plantilla' not in request.files:
            return jsonify({'error': 'No se enviaron archivos'}), 400
        
        archivo = request.files['archivo']
        plantilla = request.files['plantilla']
        
        # Verificar si los archivos tienen nombres
        if archivo.filename == '' or plantilla.filename == '':
            return jsonify({'error': 'No se seleccionaron archivos'}), 400
        
        # Verificar si los archivos son válidos
        if not (allowed_file(archivo.filename) and allowed_file(plantilla.filename)):
            return jsonify({'error': 'Formato de archivo no válido'}), 400
        
        # Guardar archivos temporalmente
        archivo_filename = secure_filename(archivo.filename)
        plantilla_filename = secure_filename(plantilla.filename)
        
        archivo_path = os.path.join(app.config['UPLOAD_FOLDER'], archivo_filename)
        plantilla_path = os.path.join(app.config['UPLOAD_FOLDER'], plantilla_filename)
        
        archivo.save(archivo_path)
        plantilla.save(plantilla_path)
        
        # Obtener carpeta de salida
        carpeta_salida = request.form.get('carpeta_salida')
        organizar_por_region = request.form.get('organizar_por_region', 'false').lower() == 'true'
        
        if not carpeta_salida:
            return jsonify({'error': 'No se especificó una carpeta de salida'}), 400
        
        # Convertir ruta relativa a absoluta si es necesario
        if not os.path.isabs(carpeta_salida):
            carpeta_salida = os.path.abspath(carpeta_salida)
        
        # Crear directorio de salida si no existe
        os.makedirs(carpeta_salida, exist_ok=True)
        
        # Leer el Excel origen
        try:
            df = pd.read_excel(archivo_path)
            print(f"Leídos {len(df)} registros del archivo origen")
            
            # Verificar si el DataFrame tiene datos
            if df.empty:
                return jsonify({'error': 'El archivo de datos está vacío'}), 400
                
            # Verificar columnas necesarias
            columnas_requeridas = ['INSTITUCION']
            columnas_opcionales = ['NOMBRES', 'APELLIDO PATERNO', 'CARRERA', 'ACTIVIDADES', 'FECHA DE INICIO', 
                                 'NOMBRE A QUIEN SE DIRIGE CARTA DE ACEPTACION', 'CARGO ESCOLAR', 'REGION']
            
            for columna in columnas_requeridas:
                if columna not in df.columns:
                    return jsonify({'error': f'No se encontró la columna requerida: {columna}'}), 400
            
            # Informar sobre columnas opcionales que faltan
            columnas_faltantes = [col for col in columnas_opcionales if col not in df.columns]
            if columnas_faltantes:
                print(f"Advertencia: No se encontraron las siguientes columnas opcionales: {', '.join(columnas_faltantes)}")
                    
        except Exception as e:
            return jsonify({'error': f'Error al leer el archivo Excel: {str(e)}'}), 500
        
        # Verificar la plantilla
        try:
            wb_plantilla = load_workbook(plantilla_path)
            print(f"Plantilla tiene {len(wb_plantilla.worksheets)} hojas")
            
            # Verificar que haya al menos 2 hojas
            if len(wb_plantilla.worksheets) < 2:
                print("Advertencia: La plantilla tiene menos de 2 hojas. Se usará la primera hoja disponible.")
                indice_hoja = 0
            else:
                indice_hoja = 1
                
            wb_plantilla.close()
            
        except Exception as e:
            return jsonify({'error': f'Error al verificar la plantilla: {str(e)}'}), 500
        
        # Procesar documentos
        try:
            if organizar_por_region and 'REGION' in df.columns:
                # Agrupar por región
                df_agrupado = df.groupby('REGION')
                
                # Crear carpeta para cada región
                for region, grupo in df_agrupado:
                    # Asegurarse de que region es un string
                    region_str = str(region) if region is not None else "sin_region"
                    ruta_region = os.path.join(carpeta_salida, region_str)
                    os.makedirs(ruta_region, exist_ok=True)
                    
                    # Procesar cada institución en la región
                    instituciones_region = grupo.groupby('INSTITUCION')
                    for institucion, datos_institucion in instituciones_region:
                        procesar_institucion(institucion, datos_institucion, plantilla_path, ruta_region, indice_hoja)
                
                return jsonify({'success': True, 'message': 'Documentos generados por región'})
            else:
                # Agrupar por institución
                instituciones = df.groupby('INSTITUCION')
                
                for institucion, datos_institucion in instituciones:
                    procesar_institucion(institucion, datos_institucion, plantilla_path, carpeta_salida, indice_hoja)
                
                return jsonify({'success': True, 'message': 'Documentos generados'})
        
        except Exception as e:
            return jsonify({'error': f'Error al procesar documentos: {str(e)}'}), 500
        
    except Exception as e:
        return jsonify({'error': f'Error: {str(e)}'}), 500

def procesar_institucion(institucion, datos_institucion, plantilla_path, ruta_salida, indice_hoja=1):
    """Procesa los datos de una institución y genera un documento"""
    # Asegurarse de que institucion es un string
    institucion_str = str(institucion) if institucion is not None else "sin_institucion"
    
    # Copiar plantilla a un archivo temporal
    nombre_archivo_temporal = f"temp_{re.sub(r'[^a-zA-Z0-9]', '_', institucion_str)}.xlsx"
    archivo_temporal = os.path.join(ruta_salida, nombre_archivo_temporal)
    shutil.copy2(plantilla_path, archivo_temporal)
    
    try:
        # Abrir el archivo temporal
        wb = load_workbook(archivo_temporal)
        
        # Verificar que el índice de la hoja sea válido
        if indice_hoja >= len(wb.worksheets):
            print(f"Advertencia: El índice de hoja {indice_hoja} no es válido. Se usará la última hoja disponible.")
            ws = wb.worksheets[-1]
        else:
            ws = wb.worksheets[indice_hoja]
        
        # Obtener la fecha de inicio más antigua
        if 'FECHA DE INICIO' in datos_institucion.columns and not datos_institucion['FECHA DE INICIO'].isna().all():
            fecha_inicio = datos_institucion.sort_values('FECHA DE INICIO').iloc[0]['FECHA DE INICIO']
            # Función para establecer valor en celda, manejando celdas fusionadas
            set_cell_value(ws, 7, 5, fecha_inicio)
        
        # Actualizar dirección (por defecto para ALTIPLANO)
        # Función para establecer valor en celda, manejando celdas fusionadas
        set_cell_value(ws, 66, 6, "Bahía de Ballenas No. 5, Piso 08, Col. Verónica Anzures, Alcaldía Miguel Hidalgo, C.P. 11300, CDMX.")
        
        # Agrupar por carrera si la columna existe
        if 'CARRERA' in datos_institucion.columns:
            carreras = datos_institucion.groupby('CARRERA')
            
            # Inicializar fila para carreras (igual que en el script PowerShell)
            fila_actual = 88
            
            # Definir estilos para las filas adicionales
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            
            # Fuente de tamaño 9
            font_size_9 = Font(size=9)
            
            # Procesar cada carrera
            for carrera, estudiantes_carrera in carreras:
                # Formatear el nombre de la carrera con solo la primera letra en mayúscula
                nombre_carrera = str(carrera) if carrera is not None else "sin_carrera"
                nombre_carrera_formateado = format_career_name(nombre_carrera)
                
                num_estudiantes = len(estudiantes_carrera)
                
                # Verificar si necesitamos formatear la fila actual
                if fila_actual > 89:  # Si estamos más allá de las filas preformateadas
                    # Aplicar formato a las celdas de la fila actual
                    for col in range(2, 9):  # Columnas B a H
                        cell = ws.cell(row=fila_actual, column=col)
                        cell.border = thin_border
                        cell.alignment = center_alignment
                        cell.font = font_size_9
                    
                    # Combinar celdas B-D, E-F, G-H para esta fila
                    ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=4)  # B-D
                    ws.merge_cells(start_row=fila_actual, start_column=5, end_row=fila_actual, end_column=6)  # E-F
                    ws.merge_cells(start_row=fila_actual, start_column=7, end_row=fila_actual, end_column=8)  # G-H
                    
                    # Ajustar el ancho de las columnas
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['E'].width = 8
                    ws.column_dimensions['G'].width = 50
                
                # Escribir nombre de carrera y número de estudiantes en la primera fila
                # Función para establecer valor en celda, manejando celdas fusionadas
                set_cell_value(ws, fila_actual, 2, nombre_carrera_formateado)
                set_cell_value(ws, fila_actual, 5, num_estudiantes)
                
                # Replicar valores en lugar de combinar celdas (igual que en el script PowerShell)
                for i in range(1, num_estudiantes):
                    fila_siguiente = fila_actual + i
                    
                    # Verificar si necesitamos formatear la fila siguiente
                    if fila_siguiente > 89:  # Si estamos más allá de las filas preformateadas
                        # Aplicar formato a las celdas de la fila siguiente
                        for col in range(2, 9):  # Columnas B a H
                            cell = ws.cell(row=fila_siguiente, column=col)
                            cell.border = thin_border
                            cell.alignment = center_alignment
                            cell.font = font_size_9
                        
                        # Combinar celdas B-D, E-F, G-H para esta fila
                        ws.merge_cells(start_row=fila_siguiente, start_column=2, end_row=fila_siguiente, end_column=4)  # B-D
                        ws.merge_cells(start_row=fila_siguiente, start_column=5, end_row=fila_siguiente, end_column=6)  # E-F
                        ws.merge_cells(start_row=fila_siguiente, start_column=7, end_row=fila_siguiente, end_column=8)  # G-H
                    
                    # Función para establecer valor en celda, manejando celdas fusionadas
                    set_cell_value(ws, fila_siguiente, 2, nombre_carrera_formateado)
                    set_cell_value(ws, fila_siguiente, 5, num_estudiantes)
                
                # Obtener actividades únicas de los estudiantes (igual que en el script PowerShell)
                actividades_unicas = []
                for _, estudiante in estudiantes_carrera.iterrows():
                    actividad = estudiante['ACTIVIDADES'] if 'ACTIVIDADES' in estudiante and pd.notna(estudiante['ACTIVIDADES']) else None
                    # Verificar si la actividad ya está en la lista
                    if actividad is not None and actividad not in actividades_unicas:
                        actividades_unicas.append(actividad)
                
                # Agregar actividades únicas (igual que en el script PowerShell)
                fila_actividad = fila_actual
                for actividad in actividades_unicas:
                    # Formatear la actividad con la función de formato de oraciones
                    actividad_formateada = format_activity_text(actividad)
                    
                    # Verificar si necesitamos formatear la fila de actividad
                    if fila_actividad > 89:  # Si estamos más allá de las filas preformateadas
                        # Aplicar formato a la celda de actividad
                        cell = ws.cell(row=fila_actividad, column=7)
                        cell.border = thin_border
                        cell.alignment = center_alignment
                        cell.font = font_size_9
                    
                    # Función para establecer valor en celda, manejando celdas fusionadas
                    set_cell_value(ws, fila_actividad, 7, actividad_formateada)
                    fila_actividad += 1
                
                # Actualizar siguiente fila (igual que en el script PowerShell)
                fila_actual = fila_actividad
        
        # Actualizar datos del responsable si están disponibles (igual que en el script PowerShell)
        if len(datos_institucion) > 0:
            responsable = datos_institucion.iloc[0]
            if 'NOMBRE A QUIEN SE DIRIGE CARTA DE ACEPTACION' in responsable.index and pd.notna(responsable['NOMBRE A QUIEN SE DIRIGE CARTA DE ACEPTACION']):
                # Función para establecer valor en celda, manejando celdas fusionadas
                set_cell_value(ws, 116, 5, responsable['INSTITUCION'])
                set_cell_value(ws, 119, 5, responsable['NOMBRE A QUIEN SE DIRIGE CARTA DE ACEPTACION'])
                set_cell_value(ws, 122, 5, responsable['CARGO ESCOLAR'])
        
        # Construir el nombre del archivo reemplazando "INST. EDUCATIVA" por el nombre de la escuela
        # Usar el mismo formato que el script PowerShell
        nombre_base_original = "2_REG. DE PROG. INST. EDUCATIVA - PEMEX 2025 ALTIPLANO.xlsx"
        nombre_limpio = re.sub(r'[^a-zA-Z0-9\s]', '', institucion_str)
        nombre_archivo = nombre_base_original.replace("INST. EDUCATIVA", nombre_limpio)
        ruta_completa = os.path.join(ruta_salida, nombre_archivo)
        
        # Guardar archivo
        wb.save(ruta_completa)
        
        # Eliminar archivo temporal
        os.remove(archivo_temporal)
        
        print(f"Documento generado: {ruta_completa}")
        
    except Exception as e:
        print(f"Error al procesar institución {institucion_str}: {str(e)}")
        
        # Eliminar archivo temporal si existe
        if os.path.exists(archivo_temporal):
            os.remove(archivo_temporal)
        
        raise e

def format_career_name(career_name):
    """
    Formatea el nombre de la carrera para que solo la primera letra de cada palabra esté en mayúscula.
    """
    # Dividir el nombre en palabras
    words = career_name.split()
    
    # Formatear cada palabra
    formatted_words = []
    for word in words:
        if word.upper() in ['DE', 'DEL', 'LA', 'LAS', 'LOS', 'Y', 'EN']:
            # Mantener estas palabras en minúsculas
            formatted_words.append(word.lower())
        else:
            # Primera letra en mayúscula, el resto en minúscula
            formatted_words.append(word.capitalize())
    
    # Unir las palabras formateadas
    return ' '.join(formatted_words)

def format_activity_text(activity_text):
    """
    Formatea el texto de la actividad para que solo la primera letra de la primera palabra esté en mayúscula,
    y la primera letra después de cada punto también esté en mayúscula.
    """
    if not activity_text:
        return activity_text
    
    # Convertir todo a minúsculas primero
    formatted_text = activity_text.lower()
    
    # Dividir el texto en oraciones usando el punto como delimitador
    sentences = re.split(r'(\.+\s*)', formatted_text)
    
    # Formatear cada oración
    formatted_sentences = []
    capitalize_next = True  # La primera oración debe comenzar con mayúscula
    
    for part in sentences:
        if part.strip() == '':
            # Espacios vacíos, mantenerlos como están
            formatted_sentences.append(part)
        elif re.match(r'\.+', part):
            # Solo puntos, mantenerlos como están
            formatted_sentences.append(part)
            capitalize_next = True  # Después de puntos, la siguiente letra debe ser mayúscula
        else:
            # Texto de la oración
            if capitalize_next:
                # Primera letra en mayúscula
                if part:
                    formatted_sentences.append(part[0].upper() + part[1:])
                else:
                    formatted_sentences.append(part)
                capitalize_next = False
            else:
                # Mantener en minúsculas
                formatted_sentences.append(part)
    
    # Unir las partes formateadas
    return ''.join(formatted_sentences)

def set_cell_value(worksheet, row, column, value):
    """
    Establece el valor de una celda, manejando celdas fusionadas.
    Si la celda está fusionada, modifica la celda superior izquierda de la fusión.
    """
    try:
        # Intentar establecer el valor directamente
        worksheet.cell(row=row, column=column).value = value
    except Exception as e:
        # Si hay un error (probablemente por celda fusionada), buscar la celda superior izquierda de la fusión
        for merged_range in worksheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            
            # Verificar si la celda está dentro del rango fusionado
            if min_row <= row <= max_row and min_col <= column <= max_col:
                # Modificar la celda superior izquierda de la fusión
                worksheet.cell(row=min_row, column=min_col).value = value
                return
        
        # Si no se encontró ninguna fusión que contenga la celda, relanzar el error original
        raise e

# Ruta para servir el frontend
@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)